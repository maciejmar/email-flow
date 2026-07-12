from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.product import Product


def import_price_list(db: Session, content: bytes, user_id: int) -> int:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active

    db.execute(delete(Product).where(Product.user_id == user_id))

    imported = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1] or row[3] is None:
            continue
        product = Product(
            user_id=user_id,
            sku=str(row[0]).strip(),
            name=str(row[1]).strip(),
            unit=str(row[2] or "szt.").strip(),
            price=Decimal(str(row[3])),
        )
        db.add(product)
        imported += 1

    db.commit()
    return imported


def find_matching_products(db: Session, user_id: int, body: str) -> list[Product]:
    products = list(db.scalars(select(Product).where(Product.user_id == user_id)))
    lowered = body.lower()
    return [
        product
        for product in products
        if product.sku.lower() in lowered or product.name.lower() in lowered
    ]

